import pytesseract
from PIL import Image
import openpyxl

# 画像からテキストを抽出
def extract_text_from_image(image_path):
    image = Image.open(image_path)
    return pytesseract.image_to_string(image)

# 画像データのファイル名を指定
image_filenames = ['test5.png', 'test6.png']  # 1つまたは2つの画像ファイル名を指定

# 群ごとの文字列を指定
groups = {
    "日本": ["大和", "敷島", "豊後", "蔵王", "淀", "吉野", "疾風", "春雲", "島風", "YAMATO", "SHIKISHIMA", "BUNGO", "ZAO", "YODO", "YOSHINO", "HAYATE", "HARUGUMO", "SHIMAKAZE"],
    "アメリカ": ["MONTANA", "OHIO", "VERMONT", "MOINES", "WORCESTER", "SALEM", "AUSTIN", "GEARING", "SHERMAN"],
    "イギリス": ["CONQUEROR", "VINCENT", "INCOMPARABLE", "MINOTAUR", "GOLIATH", "GIBRALTAR", "PLYMOUTH", "DARING", "DRUID"],
    "フランス": ["PUBLIQUE", "BOU", "HENRI", "MARSEILLE", "COLBERT", "KLEBER", "MARCEAU"],
    "イタリア": ["COLOMBO", "VENEZIA", "REGOLO"],
    "ドイツ": ["SCHLIEFFEN", "PREUSSEN", "RST", "MECK", "HINDENBURG", "ELBING", "52", "42"],
    "ソ連": ["KREMLIN", "SLA", "NEVSKY", "PETROP", "MOSKVA", "STALINGRAD", "SEVAST", "KHABARO", "DELNY", "GROZOVOI"],
    "ヨーロッパ": ["HALLAND", "GDANSK", "RAGNAR"],
    "パンアジア": ["JINAN", "YUEYANG"],
    "オランダ": ["GOUDEN", "TROMP"],
    "スペイン": ["CASTILLA", "VARO"],
    "連合": ["BRISBANE", "VAMPIRE", "MARTIN"],
    "禁止艦艇": ["ARP", "LOUISIANA", "THUNDERER", "LAURIA", "SMOLEN", "NAPOLI", "PUERTO", "LUSHUN", "SMALAND", "SOMERS", "CLR"],
    "ブラック艦艇は目視確認": ["EXAMPLE"]
}

# グループごとの出現回数をカウント
group_counts = {group: [0, 0] for group in groups}

for i, image_filename in enumerate(image_filenames):
    text = extract_text_from_image(image_filename)
    text = text.replace('\n', ' ')  # テキスト内の改行をスペースに置換
    for group, keywords in groups.items():
        for keyword in keywords:
            group_counts[group][i] += text.count(keyword)

# 結果をExcelファイルに書き込む
workbook = openpyxl.Workbook()
worksheet = workbook.active

# ヘッダーを追加
worksheet['A1'] = "Group"
worksheet['B1'] = "Alpha"
worksheet['C1'] = "Bravo"

for i, (group, counts) in enumerate(group_counts.items(), start=2):
    worksheet.cell(row=i, column=1, value=group)
    worksheet.cell(row=i, column=2, value=counts[0])
    worksheet.cell(row=i, column=3, value=counts[1])

# Excelファイルを保存
workbook.save('group_counts.xlsx')
